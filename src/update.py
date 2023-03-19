import openpyxl, csv, os

# First, backup the .xlsx
bu = openpyxl.load_workbook(filename="out/2023_NBA_Standing_Chart.xlsx")
bu.save('out/backup.xlsx')
bu.close()

# Get tonight's game stats
filename = os.listdir('dat')

game_stats = []
with open(f"dat/{filename[0]}") as f:
    reader = csv.reader(f)
    for row in reader:
        game_stats.append(row)

os.remove(f"dat/{filename[0]}")

# game_stats format:
#   [['DET', 'L', 'CHA', 'W'], ['MIA', 'W', 'PHI', 'L'], ...]

# Plug stats into the master spreadsheet
wb = openpyxl.load_workbook(filename="out/2023_NBA_Standing_Chart.xlsx")
ws = wb.active

for game in game_stats:
    #away_team = game[0]
    #away_fin = game[1]
    #home_team = game[2]
    #home_fin = game[3]

    # update team W/L
    for i in range(0,2):
        for r in range(2,32):
            if (ws.cell(row=r, column=1).value == game[2*i]):
                last_val = 0
                for c in range(2,84):
                    if (ws.cell(row=r, column=c).value != None):
                        last_val = ws.cell(row=r, column=c).value
                    else:
                        #print(f"{last_val} at ({r}, {c})")
                        print(f"{game[2*i]} {game[2*i+1]}")
                        break

                # (r, c) is the cell to be updated
                # last_val is the value at (r, c-1)

                if (game[2*i+1] == "W"):
                    ws.cell(row=r, column=c, value=last_val+1) 
                else:
                    ws.cell(row=r, column=c, value=last_val) 

# Save and close
wb.save('out/2023_NBA_Standing_Chart.xlsx')
wb.close()
